import frappe
import os
import requests
import traceback
from frappe.utils import now, now_datetime, add_days, getdate
from datetime import datetime, timedelta
import calendar
import openpyxl
from frappe.utils.file_manager import get_file_path


@frappe.whitelist(allow_guest=False)
def import_yearly_report(file_url=None):
    """
    Import file Excel xuất theo năm, với nhiều sheet tương ứng với các tháng.
    Tạo Lunch Order và Transaction cho từng tháng.
    """
    if not file_url:
        frappe.throw("Cần cung cấp file_url của file Excel")

    frappe.flags.skip_update_wallet = True
    try:
        frappe.log_error(f"[IMPORT] Bắt đầu import file: {file_url}", "Import Process")
        file_path = get_file_path(file_url)
        frappe.log_error(f"[IMPORT] File path: {file_path}", "Import Process")
        
        # 2. Mở file trực tiếp từ ổ cứng (Nhanh và không lỗi URL)
        wb = openpyxl.load_workbook(file_path, data_only=True)
        frappe.log_error(f"[IMPORT] Load workbook thành công. Sheets: {wb.sheetnames}", "Import Process")

        # Xử lý từng sheet (mỗi sheet là một tháng)
        for sheet_name in wb.sheetnames:
            if not sheet_name.startswith("Tháng "):
                frappe.log_error(f"[IMPORT] Skip sheet không phải tháng: {sheet_name}", "Import Process")
                continue

            month_part = sheet_name.replace("Tháng ", "").split("-")[0]
    
            month = int(month_part)
            ws = wb[sheet_name]
            frappe.log_error(f"[IMPORT] Processing sheet: {sheet_name}, Month: {month}", "Import Process")

            # Tìm session cho tháng này (giả sử năm hiện tại)
            year = datetime.now().year
            frappe.log_error(f"[IMPORT] Tìm session cho tháng {month}/{year}", "Import Process")
            session_name = frappe.db.get_value("Lunch Session", {"date": ["between", [datetime(year, month, 1), datetime(year, month, calendar.monthrange(year, month)[1])]]}, "name")
            if not session_name:
                frappe.log_error(f"[IMPORT] ❌ Không tìm thấy session cho tháng {month}/{year}", "Import Process")
                frappe.msgprint(f"Không tìm thấy session cho tháng {month}")
                continue
            frappe.log_error(f"[IMPORT] ✓ Tìm thấy session: {session_name}", "Import Process")

            session_doc = frappe.get_doc("Lunch Session", session_name)
            menu_items = session_doc.get("menu_items")  # Giả sử có field menu_items
            if not menu_items:
                frappe.log_error(f"[IMPORT] ❌ Session {session_name} không có menu_items", "Import Process")
                frappe.msgprint(f"Session {session_name} không có menu_items")
                continue
            frappe.log_error(f"[IMPORT] ✓ Session có {len(menu_items)} menu items", "Import Process")

            # Giả sử dùng menu_item đầu tiên
            menu_item = menu_items[0].menu_item if menu_items else None
            if not menu_item:
                frappe.log_error(f"[IMPORT] ❌ Không thể lấy menu_item từ session", "Import Process")
                continue
            frappe.log_error(f"[IMPORT] ✓ Sử dụng menu_item: {menu_item}", "Import Process")

            # Đọc dữ liệu từ sheet
            rows = list(ws.iter_rows(values_only=True))
            headers = rows[2]      # Dòng thứ 3 (index 2)
            data_rows = rows[3:]
            frappe.log_error(f"[IMPORT] ✓ Đọc {len(data_rows)} dòng dữ liệu", "Import Process")

            date_columns = []
            for i, header in enumerate(headers):
                if header and str(header).isdigit():
                    date_columns.append((i, int(header)))
            frappe.log_error(f"[IMPORT] ✓ Tìm thấy {len(date_columns)} cột ngày: {[day for _, day in date_columns]}", "Import Process")

            for row in data_rows:
                if not row[1]:  # Họ và tên
                    continue

                user_name = row[1]
                frappe.log_error(f"[IMPORT] Processing user: {user_name}", "Import Process")
                # Tìm Zalo User Map
                zalo_user = frappe.db.get_value("Zalo User Map", {"real_name": user_name}, "name")
                if not zalo_user:
                    frappe.log_error(f"[IMPORT] ⚠️ Không tìm thấy Zalo User cho: {user_name}", "Import Process")
                    continue
                frappe.log_error(f"[IMPORT] ✓ User {user_name} -> Zalo User: {zalo_user}", "Import Process")

                # Tạo order cho những ngày có 1
                for col_idx, day in date_columns:
                    if str(row[col_idx]).strip() == "1":
                        order_date = datetime(year, month, day)
                        frappe.log_error(f"[IMPORT]   Ngày {day}: Có order", "Import Process")
                        # Kiểm tra đã có order chưa
                        existing = frappe.db.exists("Lunch Order", {
                            "zalo_user": zalo_user,
                            "session": session_name,
                            "created_at": ["between", [order_date, order_date + timedelta(days=1) - timedelta(seconds=1)]]
                        })
                        if existing:
                            frappe.log_error(f"[IMPORT]   ⚠️ Order đã tồn tại cho ngày {day}", "Import Process")
                            continue

                        # Tạo Lunch Order
                        try:
                            order = frappe.get_doc({
                                "doctype": "Lunch Order",
                                "zalo_user": zalo_user,
                                "session": session_name,
                                "menu_item": menu_item,
                                "created_at": order_date,
                                "is_active": 1
                            })
                            order.insert(ignore_permissions=True)
                            frappe.log_error(f"[IMPORT]   ✓ Created Order: {order.name}", "Import Process")
                        except Exception as e:
                            frappe.log_error(f"[IMPORT]   ❌ Lỗi tạo Order: {str(e)}", "Import Process")
                            continue

                        # Tạo Transaction (pay)
                        try:
                            menu_price = frappe.db.get_value("Lunch Menu Item", menu_item, "price") or 0
                            transaction = frappe.get_doc({
                                "doctype": "Transaction",
                                "zalo_user": zalo_user,
                                "type": "Pay",
                                "amount": -abs(menu_price),
                                "reference": order.name,
                                "date": order_date,
                                "description": f"Thanh toán cho bữa ăn ngày {order_date.strftime('%d/%m/%Y')}"
                            })
                            transaction.insert(ignore_permissions=True)
                            frappe.log_error(f"[IMPORT]   ✓ Created Transaction: {transaction.name} (Amount: {-abs(menu_price)})", "Import Process")
                        except Exception as e:
                            frappe.log_error(f"[IMPORT]   ❌ Lỗi tạo Transaction: {str(e)}", "Import Process")

        frappe.log_error(f"[IMPORT] ✓✓✓ Import hoàn tất thành công", "Import Process")
        frappe.msgprint("Import thành công")

    except Exception as e:
        frappe.log_error(f"[IMPORT] ❌ Import error: {str(e)}\n{traceback.format_exc()}", "Import Process")
        frappe.throw(f"Lỗi import: {str(e)}")

    finally:
        frappe.flags.skip_update_wallet = False
