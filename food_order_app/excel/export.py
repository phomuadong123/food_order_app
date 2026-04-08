import frappe
import os
import requests
import traceback
from frappe.utils import now, nowdate, add_days, getdate
from datetime import date, datetime, timedelta
import calendar


def _to_datetime_start(date_obj):
    return datetime(date_obj.year, date_obj.month, date_obj.day, 0, 0, 0)


def _to_datetime_end(date_obj):
    return datetime(date_obj.year, date_obj.month, date_obj.day, 23, 59, 59)


def _get_transaction_maps(start_date, end_date):
    deposits = frappe.db.sql(
        """
        SELECT
            t.zalo_user,
            SUM(t.amount) AS deposit_amount
        FROM `tabTransaction` t
        WHERE t.type = 'Deposit'
            AND (t.reference IS NULL OR t.reference = '')
            AND t.date >= %s
            AND t.date < DATE_ADD(%s, INTERVAL 1 MONTH)
        GROUP BY t.zalo_user
        """,
        (start_date, start_date),
        as_dict=True,
    )
    deposit_map = {d.zalo_user: float(d.deposit_amount or 0) for d in deposits}

    sum_in_period = frappe.db.sql(
        """
        SELECT
            t.zalo_user,
            SUM(t.amount) AS sum_amount
        FROM `tabTransaction` t
        WHERE 
            t.date <= LAST_DAY(%s - INTERVAL 1 MONTH) 
            AND t.type = 'Pay'
        GROUP BY 
            t.zalo_user;
        """,
        (start_date),
        as_dict=True,
    )
    sum_in_period_map = {d.zalo_user: float(d.sum_amount or 0) for d in sum_in_period}

    sum_after_end = frappe.db.sql(
        """
        SELECT
                t.zalo_user,
                SUM(t.amount) AS sum_amount
            FROM `tabTransaction` t
            WHERE t.type = 'Deposit'
                AND (t.reference IS NULL OR t.reference = '')
                AND t.date < %s 
            GROUP BY t.zalo_user;
        """,
        (start_date,),  # 
        as_dict=True,
    )
    sum_after_end_map = {d.zalo_user: float(d.sum_amount or 0) for d in sum_after_end}

    return deposit_map, sum_in_period_map, sum_after_end_map


def _create_report_sheet(wb, start_date, end_date, date_headers, period_query, sheet_title):
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    orders = frappe.db.sql(
        f"""
        SELECT
            lo.zalo_user,
                DAY(
                CASE 
                    WHEN TIME(lo.created_at) > '12:00:00' 
                    THEN DATE_ADD(lo.created_at, INTERVAL 1 DAY)
                    ELSE lo.created_at
                END
            ) AS period_index,
            lmi.price
        FROM `tabLunch Order` lo
        JOIN `tabLunch Session` ls ON lo.session = ls.name
        JOIN `tabLunch Menu Item` lmi ON lo.menu_item = lmi.name
        WHERE ls.date BETWEEN %s AND %s
            AND ls.status != 'Draft' AND lo.is_active = 1
        """,
        (start_date, end_date),
        as_dict=True,
    )

    is_future = getdate(start_date) > getdate(nowdate())

    user_orders = {}
    for o in orders:
        if not o.zalo_user:
            continue
        period_index = int(o.period_index or 0)
        if period_index < 1 or period_index > len(date_headers):
            continue

        if o.zalo_user not in user_orders:
            user_orders[o.zalo_user] = {'days': set(), 'total_amount': 0}

        user_orders[o.zalo_user]['days'].add(period_index)
        user_orders[o.zalo_user]['total_amount'] += (o.price or 0)

   
    wallets = frappe.get_all("Lunch Wallet", fields=["zalo_user", "balance"])
    wallet_map = {w.zalo_user: float(w.balance or 0) for w in wallets}

    deposit_map, sum_in_period_map, sum_after_end_map = _get_transaction_maps(start_date, end_date)
    frappe.log_error(
        title="Important Debug Info", 
        message=str(sum_in_period_map)
    )
    frappe.log_error(
        title="Important Debug Info", 
        message=str(sum_after_end_map)
    )

    if len(wb.sheetnames) == 1 and wb.active.title == 'Sheet' and wb.active.max_row == 1 and wb.active['A1'].value is None:
        ws = wb.active
        ws.title = sheet_title
    else:
        ws = wb.create_sheet(title=sheet_title)

    header_fill = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")
    header_font = Font(bold=True)
    center_align = Alignment(horizontal='center', vertical='center')

    ws.merge_cells(start_row=1, start_column=3, end_row=1, end_column=2 + len(date_headers))
    day_cell = ws.cell(row=1, column=3, value="Ngày")
    day_cell.alignment = center_align
    day_cell.font = header_font
    day_cell.fill = header_fill

    num_columns = 2 + len(date_headers) + 6
    for col in range(1, num_columns + 1):
        c = ws.cell(row=1, column=col)
        c.fill = header_fill

    summary_headers = [
        "Số ngày ăn",
        "Đơn giá suất ăn \n(VNĐ)",
        "Thành tiền \n(VNĐ)",
        "Số tiền đầu kỳ \n(VNĐ)",
        "Số tiền nạp vào \n(VNĐ)",
        "Số tiền còn lại \n(VNĐ)",
    ]
    headers = ["STT", "Họ và tên"] + date_headers + summary_headers
    ws.append(headers)

    for col_idx, _ in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col_idx)
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.fill = header_fill

        if col_idx == 1:
            ws.column_dimensions[get_column_letter(col_idx)].width = 5
        elif col_idx == 2:
            ws.column_dimensions[get_column_letter(col_idx)].width = 25
        elif 2 < col_idx <= 2 + len(date_headers):
            ws.column_dimensions[get_column_letter(col_idx)].width = 4
        else:
            ws.column_dimensions[get_column_letter(col_idx)].width = 15

    users = frappe.get_all("Zalo User Map", fields=["name", "real_name"])
    stt = 1
    for u in users:
        u_data = user_orders.get(u.name, {'days': set(), 'total_amount': 0})
        num_days = len(u_data['days'])
        total_price = u_data['total_amount']
        avg_price = (total_price / num_days) if num_days > 0 else 0
        current_balance = wallet_map.get(u.name, 0)
        sum_in_period = sum_in_period_map.get(u.name, 0)
        sum_after_end = sum_after_end_map.get(u.name, 0)
        deposit_amount = deposit_map.get(u.name, 0)

        if is_future:
            beginning_balance = 0
            end_balance = 0
        else:
            beginning_balance = sum_in_period + sum_after_end
            end_balance = beginning_balance - total_price + deposit_amount

        row_data = [stt, u.real_name or u.full_name]
        for idx in range(1, len(date_headers) + 1):
            row_data.append(1 if idx in u_data['days'] else "")

        row_data.extend([
            num_days,
            avg_price,
            total_price,
            beginning_balance,
            deposit_amount,
            end_balance,
        ])

        ws.append(row_data)

        curr_row = ws.max_row
        for col_idx in range(1, len(row_data) + 1):
            cell = ws.cell(row=curr_row, column=col_idx)
            if col_idx > 2:
                cell.alignment = center_align

        currency_start = 2 + len(date_headers) + 2
        currency_end = 2 + len(date_headers) + 6
        for col_idx in range(currency_start, currency_end + 1):
            ws.cell(row=curr_row, column=col_idx).number_format = '#,##0'

        stt += 1

    ws.freeze_panes = "C3"


def _build_excel_report(start_date, end_date, date_headers, period_query, title, filename_suffix):
    import openpyxl
    from io import BytesIO

    wb = openpyxl.Workbook()
    _create_report_sheet(wb, start_date, end_date, date_headers, period_query, title)

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    frappe.response['filename'] = f"BaoCao_AnTrua_{filename_suffix}.xlsx"
    frappe.response['filecontent'] = output.getvalue()
    frappe.response['type'] = 'binary'


@frappe.whitelist(allow_guest=False)
def export_monthly_report(month=None, year=None):
    if not month or not year:
        current = datetime.now()
        month = current.month
        year = current.year

    month = int(month)
    year = int(year)
    _, days_in_month = calendar.monthrange(year, month)

    start_date = date(year, month, 1)
    end_date = date(year, month, days_in_month)
    date_headers = [str(day) for day in range(1, days_in_month + 1)]
    title = f"Tháng {month}-{year}"
    filename_suffix = f"Thang_{month}_{year}"
    return _build_excel_report(start_date, end_date, date_headers, "DAY(lo.created_at)", title, filename_suffix)


@frappe.whitelist(allow_guest=False)
def export_daily_report(date=None):
    if not date:
        date = datetime.now().date()
    else:
        date = getdate(date)

    start_date = _to_datetime_start(date)
    end_date = _to_datetime_end(date)
    date_headers = [date.strftime('%d/%m/%Y')]
    title = f"Ngày {date.strftime('%d-%m-%Y')}"
    filename_suffix = f"Ngay_{date.strftime('%Y%m%d')}"
    return _build_excel_report(start_date, end_date, date_headers, '1', title, filename_suffix)


@frappe.whitelist(allow_guest=False)
def export_yearly_report(year=None):
    import openpyxl
    from io import BytesIO

    if not year:
        year = datetime.now().year

    year = int(year)
    wb = openpyxl.Workbook()

    for month in range(1, 13):
        _, days_in_month = calendar.monthrange(year, month)
        start_date = datetime(year, month, 1, 0, 0, 0)
        end_date = datetime(year, month, days_in_month, 23, 59, 59)
        date_headers = [str(day) for day in range(1, days_in_month + 1)]
        sheet_title = f"Thang_{month}"
        _create_report_sheet(wb, start_date, end_date, date_headers, "DAY(ls.date)", sheet_title)

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    frappe.response['filename'] = f"BaoCao_AnTrua_Nam_{year}.xlsx"
    frappe.response['filecontent'] = output.getvalue()
    frappe.response['type'] = 'binary'
