import frappe
import os
import requests
import traceback
from frappe.utils import now, now_datetime, add_days, getdate
from datetime import datetime, timedelta

logger = frappe.logger("lunch_api")

# =========================
# ZALO CONFIG
# =========================

ZALO_APP_ID = os.getenv("ZALO_APP_ID")
ZALO_SECRET = os.getenv("ZALO_SECRET")
REDIRECT_URI = os.getenv("ZALO_REDIRECT_URI")
ZALO_OA_ACCESS_TOKEN = os.getenv("ZALO_OA_ACCESS_TOKEN")
GROUP_ID_ZALO = os.getenv("GROUP_ID_ZALO")
BASE_URL = os.getenv("BASE_URL")


# =========================
# START VOTE
# =========================

@frappe.whitelist(allow_guest=True)
def start_vote(session):
    if not ZALO_APP_ID:
        logger.warning("ZALO_APP_ID is not configured")
        return {"error": "start_vote_failed", "detail": "ZALO_APP_ID not configured"}

    if not REDIRECT_URI:
        logger.warning("REDIRECT_URI is not configured")
        return {"error": "start_vote_failed", "detail": "REDIRECT_URI not configured"}

    if not session:
        logger.warning("Session is missing")
        return {"error": "start_vote_failed", "detail": "Missing session"}

    try:
        base = BASE_URL or frappe.utils.get_url()
        redirect_uri = f"{base}{REDIRECT_URI}"

        # encode redirect URI to be safe
        from urllib.parse import quote_plus
        encoded_redirect_uri = quote_plus(redirect_uri)

        oauth_url = (
            f"https://oauth.zaloapp.com/v4/permission?"
            f"app_id={ZALO_APP_ID}"
            f"&redirect_uri={encoded_redirect_uri}"
            f"&state={session}"
        )

        # Logger chỉ ghi ngắn gọn để tránh Value too big
        logger.info(f"ZALO OAuth URL generated for session {session}")

        frappe.local.response["type"] = "redirect"
        frappe.local.response["location"] = oauth_url
        return

    except Exception as err:
        err_trace = frappe.get_traceback()
        logger.error(f"start_vote failed: {err}")
        logger.debug(err_trace)
        return {"error": "start_vote_failed", "detail": str(err)}


# =========================
# ZALO CALLBACK
# =========================

@frappe.whitelist(allow_guest=True)
def zalo_callback(code=None, state=None):

    logger.info(f"ZALO CALLBACK START (code={'<redacted>' if code else None}, state={state})")

    try:

        if not code:
            logger.warning("Missing OAuth code")
            return {"error": "missing_code"}

        # =====================================================
        # STEP 1: GET ACCESS TOKEN
        # =====================================================

        try:

            token_res = requests.post(
                "https://oauth.zaloapp.com/v4/access_token",
                headers={
                    "Content-Type": "application/x-www-form-urlencoded",
                    "secret_key": ZALO_SECRET
                },
                data={
                    "app_id": ZALO_APP_ID,
                    "code": code,
                    "grant_type": "authorization_code"
                },
                timeout=15
            )

            logger.info(f"ZALO TOKEN RESPONSE status={token_res.status_code}")

            token_json = token_res.json()

        except Exception as e:
            err_trace = frappe.get_traceback()
            logger.error(f"ZALO TOKEN API FAILED: {e}")
            logger.debug(err_trace)
            return {"error": "token_api_failed", "detail": str(e)}

        access_token = token_json.get("access_token")

        if not access_token:
            logger.warning("TOKEN NOT FOUND from zalo response")
            return {"error": "token_failed", "detail": "access_token missing"}

        logger.info("ACCESS TOKEN RECEIVED")

        # =====================================================
        # STEP 2: GET USER PROFILE
        # =====================================================

        try:

            profile_res = requests.get(
                "https://graph.zalo.me/v2.0/me",
                params={
                    "fields": "id,name,picture",
                    "access_token": access_token
                },
                timeout=15
            )

            logger.info(f"ZALO PROFILE RESPONSE status={profile_res.status_code}")

            profile = profile_res.json()

        except Exception as e:
            err_trace = frappe.get_traceback()
            logger.error(f"PROFILE API FAILED: {e}")
            logger.debug(err_trace)
            return {"error": "profile_api_failed", "detail": str(e)}

        zalo_id = profile.get("id")
        name = profile.get("name")

        picture = profile.get("picture", {})

        if isinstance(picture, dict):
            avatar = picture.get("data", {}).get("url", "")
        else:
            avatar = picture

        logger.info(f"PROFILE PARSED (id={zalo_id}, name={name})")

        if not zalo_id:
            logger.warning("PROFILE MISSING ID")
            return {"error": "profile_failed"}

        # =====================================================
        # STEP 3: FIND USER
        # =====================================================

        try:

            user = frappe.db.exists(
                "Zalo User Map",
                {"zalo_id": zalo_id}
            )

            logger.info(f"USER EXISTS RESULT: {user}")
            logger.info(f"CURRENT SITE: {frappe.local.site}")

        except Exception as e:
            err_trace = frappe.get_traceback()
            logger.error(f"USER LOOKUP FAILED: {e}")
            logger.debug(err_trace)
            return {"error": "user_lookup_failed", "detail": str(e)}

        # =====================================================
        # STEP 4: CREATE USER
        # =====================================================

        if not user:

            try:

                user_doc = frappe.get_doc({
                    "doctype": "Zalo User Map",
                    "zalo_id": zalo_id,
                    "full_name": name,
                    "avatar": avatar,
                    "is_active": 0,
                    "created_at": now()
                })

                user_doc.insert(ignore_permissions=True)

                frappe.db.commit()

                logger.info(f"USER INSERTED: {user_doc.name}")

                user = user_doc.name

            except Exception as e:
                err_trace = frappe.get_traceback()
                logger.error(f"USER INSERT FAILED: {e}")
                logger.debug(err_trace)
                return {"error": "user_create_failed", "detail": str(e)}

            # =====================================================
            # STEP 5: CREATE WALLET
            # =====================================================

            try:

                wallet_doc = frappe.get_doc({
                    "doctype": "Lunch Wallet",
                    "zalo_user": user,
                    "balance": 0,
                    "updated_at": now()
                })

                wallet_doc.insert(ignore_permissions=True)

                frappe.db.commit()

                logger.info(f"WALLET CREATED: {wallet_doc.name}")

            except Exception as e:

                err_trace = frappe.get_traceback()
                logger.error(f"WALLET CREATE FAILED: {e}")
                logger.debug(err_trace)

        # =====================================================
        # STEP 6: REDIRECT
        # =====================================================

        vote_page = f"/vote?session={state}&zalo_id={zalo_id}"

        logger.info(f"REDIRECT TO VOTE PAGE {vote_page}")

        frappe.local.response["type"] = "redirect"
        frappe.local.response["location"] = vote_page

        return

    except Exception as e:

        err_trace = frappe.get_traceback()
        logger.error(f"ZALO CALLBACK FATAL ERROR: {e}")
        logger.debug(err_trace)

        return {
            "error": "callback_crash",
            "trace": err_trace
        }


# =========================
# UPDATE SESSION STATS
# =========================

def update_session_stats(session):

    try:

        result = frappe.db.sql(
            """
            SELECT COUNT(*), SUM(price)
            FROM `tabLunch Order`
            WHERE session=%s
            """,
            session
        )[0]

        total_orders = result[0] or 0
        total_amount = result[1] or 0

        frappe.db.set_value(
            "Lunch Session",
            session,
            {
                "total_orders": total_orders,
                "total_amount": total_amount
            }
        )

        logger.info(f"{session} orders={total_orders} amount={total_amount}")

    except Exception as e:

        err_trace = frappe.get_traceback()
        logger.error(f"SESSION STATS ERROR: {e}")
        logger.debug(err_trace)


# =========================
# CREATE VOTE LINK
# =========================

def create_vote_link(doc, method):

    try:

        base = frappe.utils.get_url()

        link = f"{base}/api/method/food_order_app.api.start_vote?session={doc.name}"

        frappe.db.set_value(
            "Lunch Session",
            doc.name,
            "vote_link",
            link
        )

        logger.info(f"VOTE LINK CREATED: {link}")

    except Exception as e:

        err_trace = frappe.get_traceback()
        logger.error(f"CREATE VOTE LINK ERROR: {e}")
        logger.debug(err_trace)


def update_session_menu_items(doc, method=None):
    try:
        frappe.db.delete("Lunch Session Menu", {"parent": doc.name})

        if hasattr(doc, 'menu_items') and doc.menu_items:
            for row in doc.menu_items:
                if row.menu_item:
                    new_link = frappe.get_doc({
                        "doctype": "Lunch Session Menu",
                        "parent": doc.name,
                        "parenttype": "Lunch Session",
                        "parentfield": "menu_items",
                        "menu_item": row.menu_item
                    })
                    new_link.insert(ignore_permissions=True)
        
        frappe.db.commit()

    except Exception as e:
        err_trace = frappe.get_traceback()
        logger.error(f"Lỗi cập nhật danh sách Menu: {e}")
        logger.debug(err_trace)


# =========================
# MY NOTIFICATION API
# =========================

@frappe.whitelist()
def my_notification_api():
    """API that displays a notification message"""
    frappe.msgprint("Thông báo từ task được gọi mỗi phút!")
    return {"status": "success", "message": "Notification displayed"}
		
@frappe.whitelist(allow_guest=True)
def get_menu(session):

    try:

        logger.info(f"get_menu called session={session}")

        if not session:
            frappe.throw("Session is required")

        items = frappe.db.sql("""
            SELECT
                mi.name,
                mi.item_name,
                mi.price
            FROM `tabLunch Session Menu` sm
            JOIN `tabLunch Menu Item` mi
                ON sm.menu_item = mi.name
            WHERE sm.parent = %s
        """, (session,), as_dict=True)

        logger.info(f"menu count={len(items)}")

        return {
            "success": True,
            "data": items
        }

    except Exception:

        error = traceback.format_exc()

        logger.error(error)

        frappe.log_error(
            title="get_menu error",
            message=error
        )

        return {
            "success": False,
            "message": "Failed to get menu"
        }

@frappe.whitelist(allow_guest=True)
def vote(session, menu_item, zalo_id):

    try:
        logger.info(f"[VOTE] session={session}, menu_item={menu_item}, zalo_id={zalo_id}")

        # ====================================
        # 1. VALIDATE INPUT
        # ====================================
        if not session or not menu_item or not zalo_id:
            return {"success": False, "message": "Thiếu thông tin cần thiết"}

        # ====================================
        # 2. GET SESSION
        # ====================================
        try:
            session_doc = frappe.get_doc("Lunch Session", session)
        except Exception:
            logger.error(f"[VOTE] Session not found: {session}")
            return {"success": False, "message": "Không tìm thấy phiên đăng ký bữa ăn"}

        if session_doc.status != "Open":
            return {"success": False, "message": "Phiên bữa ăn đã đóng, không thể đăng ký"}

        # ====================================
        # 3. GET USER
        # ====================================
        user = frappe.db.get_value("Zalo User Map", {"zalo_id": zalo_id}, "name")
        if not user:
            return {"success": False, "message": "Người dùng không tồn tại"}

        is_active = frappe.db.get_value("Zalo User Map", user, "is_active")
        if not int(is_active or 0):
            return {"success": False, "message": "Vui lòng đợi admin kích hoạt tài khoản"}

        # ====================================
        # 4. CHECK DUPLICATE ORDER (only active)
        # ====================================
        existed = frappe.db.exists("Lunch Order", {"session": session, "zalo_user": user, "is_active": 1})
        if existed:
            return {"success": False, "message": "Người dùng đã đăng ký bữa ăn cho phiên này rồi"}

        # ====================================
        # 5. GET PRICE (JOIN Lunch Session Menu + Lunch Menu Item)
        # ====================================
        row = frappe.db.sql("""
            SELECT mi.price
            FROM `tabLunch Session Menu` sm
            JOIN `tabLunch Menu Item` mi
                ON sm.menu_item = mi.name
            WHERE sm.parent = %s AND sm.menu_item = %s
            LIMIT 1
        """, (session, menu_item), as_dict=True)

        if not row:
            return {"success": False, "message": "Danh sách đăng ký bữa ăn không có sẵn hôm nay"}

        price = row[0].price
        logger.info(f"[VOTE] PRICE={price}")

        # ====================================
        # 6. GET WALLET
        # ====================================
        wallet_name = frappe.db.get_value("Lunch Wallet", {"zalo_user": user}, "name")
        wallet_balance = frappe.db.get_value("Lunch Wallet", wallet_name, "balance")

        if wallet_balance is None:
            return {"success": False, "message": "không tìm thấy ví của người dùng"}

        logger.info(f"[VOTE] WALLET balance={wallet_balance}")

        # ====================================
        # 7. INSERT ORDER
        # ====================================
        try:
            order_doc = frappe.get_doc({
                "doctype": "Lunch Order",
                "session": session,
                "zalo_user": user,
                "is_active": 1,
                "menu_item": menu_item,
                "created_at": now()
            })

            order_doc.insert(ignore_permissions=True)

            transaction = frappe.get_doc({
                "doctype": "Transaction",
                "zalo_user": user,
                "type": "Pay",
                "amount": -price,  # Trừ tiền nên số âm
                "reference": order_doc.name, # Ghi nhận order vừa tạo
                "session": session,
                "description": "Trừ tiền cho order",
                "date": now()
            })

            transaction.insert(ignore_permissions=True)

            frappe.db.commit()
        except Exception:
            frappe.db.rollback()
            logger.error(traceback.format_exc())
            return {"success": False, "message": "Tạo order đăng ký bữa ăn thất bại"}

        logger.info(f"[VOTE] ORDER CREATED name={order_doc.name}")

        # ====================================
        # 9. SUCCESS
        # ====================================
        return {"success": True, "message": "Đăng ký bữa ăn thành công", "order": order_doc.name}

    except Exception:
        frappe.db.rollback()
        error = traceback.format_exc()
        logger.error(f"[VOTE] ERROR {error}")
        return {"success": False, "message": "Lỗi khi đăng ký bữa ăn"}

@frappe.whitelist(allow_guest=True)
def cancel_vote(session, zalo_id):
    try:
        if not session or not zalo_id:
            return {"success": False, "message": "Missing parameters"}

        user = frappe.db.get_value("Zalo User Map", {"zalo_id": zalo_id}, "name")
        if not user:
            return {"success": False, "message": "Người dùng không tồn tại"}

        order_name = frappe.db.get_value("Lunch Order", {"session": session, "zalo_user": user, "is_active": 1}, "name")
        if not order_name:
            return {"success": False, "message": "Active order not found"}

        order_doc = frappe.get_doc("Lunch Order", order_name)

        # Get price of menu item for this session
        row = frappe.db.sql("""
            SELECT mi.price
            FROM `tabLunch Session Menu` sm
            JOIN `tabLunch Menu Item` mi
                ON sm.menu_item = mi.name
            WHERE sm.parent = %s AND sm.menu_item = %s
            LIMIT 1
        """, (session, order_doc.menu_item), as_dict=True)

        if not row:
            return {"success": False, "message": "Danh sách đăng ký bữa ăn không có sẵn hôm nay"}

        price = row[0].price

        order_doc.is_active = 0
        order_doc.save(ignore_permissions=True)

        refund_tx = frappe.get_doc({
            "doctype": "Transaction",
            "zalo_user": user,
            "type": "Deposit",
            "amount": price,
            "reference": order_doc.name,
            "session": session,
            "description": "Hoàn tiền khi hủy đăng ký bữa ăn",
            "date": now()
        })

        refund_tx.insert(ignore_permissions=True)

        frappe.db.commit()

        return {"success": True, "message": "Hủy đăng ký bữa ăn thành công"}

    except Exception:
        frappe.db.rollback()
        error = traceback.format_exc()
        logger.error(f"[CANCEL] ERROR {error}")
        return {"success": False, "message": "Internal error"}

@frappe.whitelist(allow_guest=True)
def get_order_status(session, zalo_id):
    try:
        if not session or not zalo_id:
            return {"success": False, "message": "Missing parameters"}

        user = frappe.db.get_value("Zalo User Map", {"zalo_id": zalo_id}, "name")
        if not user:
            return {"success": False, "message": "Người dùng không tồn tại"}

        session_date = frappe.db.get_value("Lunch Session", session, "date")
        if not session_date:
            return {"success": False, "message": "Phiên đăng ký bữa ăn không tồn tại"}

        has_order = frappe.db.exists("Lunch Order", {"session": session, "zalo_user": user, "is_active": 1})

        return {
            "success": True,
            "has_order": bool(has_order),
            "date": str(session_date)
        }

    except Exception:
        error = traceback.format_exc()
        logger.error(f"[GET_ORDER_STATUS] ERROR {error}")
        return {"success": False, "message": "Internal error"}

@frappe.whitelist(allow_guest=True)
def get_user_activation_status(zalo_id):
    try:
        if not zalo_id:
            return {"success": False, "message": "Missing zalo_id"}

        row = frappe.db.get_value(
            "Zalo User Map",
            {"zalo_id": zalo_id},
            ["name", "is_active"],
            as_dict=True
        )
        if not row:
            return {"success": False, "message": "User not found"}

        is_active = bool(int(row.is_active or 0))
        return {
            "success": True,
            "is_active": is_active,
            "message": "OK" if is_active else "đợi admin active"
        }
    except Exception:
        error = traceback.format_exc()
        logger.error(f"[GET_USER_ACTIVATION_STATUS] ERROR {error}")
        return {"success": False, "message": "Internal error"}

@frappe.whitelist(allow_guest=True)
def get_support_group_info():
    try:
        row = frappe.get_all(
            "Zalo Group",
            fields=["group_id", "group_link", "modified"],
            order_by="modified desc",
            limit=1
        )

        if not row:
            return {"success": True, "data": None}

        return {"success": True, "data": row[0]}
    except Exception:
        error = traceback.format_exc()
        logger.error(f"[GET_SUPPORT_GROUP_INFO] ERROR {error}")
        return {"success": False, "message": "Internal error"}

@frappe.whitelist(allow_guest=True)
def get_session_votes(session):
    try:
        if not session:
            return {"success": False, "message": "Missing session"}

        rows = frappe.db.sql("""
            SELECT
                lo.name AS order_id,
                zum.zalo_id,
                COALESCE(NULLIF(zum.real_name, ''), NULLIF(zum.full_name, ''), zum.zalo_id) AS voter_name,
                lmi.item_name AS menu_item_name,
                lmi.price,
                lo.created_at
            FROM `tabLunch Order` lo
            LEFT JOIN `tabZalo User Map` zum
                ON lo.zalo_user = zum.name
            LEFT JOIN `tabLunch Menu Item` lmi
                ON lo.menu_item = lmi.name
            WHERE lo.session = %s
                AND lo.is_active = 1
            ORDER BY lo.created_at DESC, lo.creation DESC
        """, (session,), as_dict=True)

        return {"success": True, "data": rows}
    except Exception:
        error = traceback.format_exc()
        logger.error(f"[GET_SESSION_VOTES] ERROR {error}")
        return {"success": False, "message": "Internal error"}


@frappe.whitelist(allow_guest=True)
def get_my_session_transactions(session, zalo_id):
    """
    Lịch sử giao dịch của user hiện tại trong session (tab Transaction).
    Số dư còn lại: tính sau mỗi giao dịch trong session (theo thứ tự thời gian).
    """
    try:
        if not session or not zalo_id:
            return {"success": False, "message": "Thiếu thông tin tham số"}

        user = frappe.db.get_value("Zalo User Map", {"zalo_id": zalo_id}, "name")
        if not user:
            return {"success": False, "message": "Người dùng không tồn tại"}

        voter_name = frappe.db.get_value(
            "Zalo User Map",
            user,
            ["full_name", "real_name", "zalo_id"],
            as_dict=True,
        )
        display_name = (
            (voter_name.get("real_name") or "").strip()
            or (voter_name.get("full_name") or "").strip()
            or voter_name.get("zalo_id")
            or user
        )

        wallet_name = frappe.db.get_value("Lunch Wallet", {"zalo_user": user}, "name")
        wallet_balance = frappe.db.get_value("Lunch Wallet", wallet_name, "balance")
        if wallet_balance is None:
            wallet_balance = 0

        rows = frappe.db.sql(
            """
            SELECT
                t.name AS transaction_id,
                t.type,
                t.amount,
                t.reference,
                t.date,
                t.description,
                COALESCE(NULLIF(lmi.item_name, ''), '') AS menu_item_name,
                COALESCE(lmi.price, ABS(t.amount)) AS menu_price
            FROM `tabTransaction` t
            LEFT JOIN `tabLunch Order` lo ON t.reference = lo.name
            LEFT JOIN `tabLunch Menu Item` lmi ON lo.menu_item = lmi.name
            WHERE t.zalo_user = %s AND t.session = %s
            ORDER BY t.date ASC, t.creation ASC
            """,
            (user, session),
            as_dict=True,
        )

        sum_session = sum(float(r.get("amount") or 0) for r in rows)
        running = float(wallet_balance) - sum_session

        out = []
        for idx, r in enumerate(rows, start=1):
            amt = float(r.get("amount") or 0)
            running += amt
            price = r.get("menu_price")
            if price is None:
                price = abs(amt) if amt else 0
            info_parts = [
                r.get("type") or "",
                str(r.get("transaction_id") or ""),
                frappe.utils.format_datetime(r.get("date")) if r.get("date") else "",
                (r.get("description") or "").strip(),
            ]
            transaction_info = " | ".join(p for p in info_parts if p)

            out.append(
                {
                    "stt": idx,
                    "voter_name": display_name,
                    "menu_item_name": r.get("menu_item_name") or "—",
                    "price": float(price or 0),
                    "balance_after": running,
                    "transaction_info": transaction_info,
                }
            )

        return {
            "success": True,
            "data": out,
            "voter_name": display_name,
            "wallet_balance": float(wallet_balance),
        }
    except Exception:
        error = traceback.format_exc()
        logger.error(f"[GET_MY_SESSION_TRANSACTIONS] ERROR {error}")
        return {"success": False, "message": "Internal error"}


@frappe.whitelist()
def update_wallet_on_transaction(doc, method=None):
    wallet = frappe.get_doc("Lunch Wallet", {"zalo_user": doc.zalo_user})
    
    if not wallet:
        frappe.throw(f"Không tìm thấy ví cho người dùng {doc.zalo_user}")
    
    new_balance = wallet.balance + doc.amount

    wallet.balance = new_balance
    wallet.updated_at = frappe.utils.now_datetime()
    wallet.save(ignore_permissions=True)

    frappe.msgprint(f"Ví của {doc.full_name} đã được cập nhật. Số dư: {new_balance:,.0f}đ")	

# =========================
# AUTO SESSION DAILY RENEWAL
# =========================


def refresh_zalo_tokens():
    try:
        # 1. Lấy thông tin từ bảng tabZalo Config (Lấy bản ghi đầu tiên)
        # Vì đây là bảng cấu hình, ta thường chỉ có 1 dòng duy nhất
        config = frappe.db.sql("""
            SELECT name, app_id, secret_key, refresh_token 
            FROM `tabZalo Config` 
            LIMIT 1
        """, as_dict=True)

        if not config:
            frappe.log_error("Bảng tabZalo Config trống rỗng. Hãy tạo 1 bản ghi trước!", "Zalo SQL Error")
            return None
        
        config = config[0]
        doc_name = config["name"] # Tên bản ghi để dùng cho WHERE

        # 2. Gọi API Zalo
        url = "https://oauth.zaloapp.com/v4/oa/access_token"
        headers = {
            'Content-Type': 'application/x-www-form-urlencoded',
            'secret_key': config["secret_key"]
        }
        payload = {
            'refresh_token': config["refresh_token"],
            'app_id': config["app_id"],
            'grant_type': 'refresh_token'
        }

        response = requests.post(url, headers=headers, data=payload)
        res_data = response.json()

        if "access_token" in res_data:
            new_at = res_data["access_token"]
            new_rt = res_data["refresh_token"]

            # 3. UPDATE theo phong cách SQL bảng thường như bạn muốn
            frappe.db.sql("""
                UPDATE `tabZalo Config`
                SET 
                    access_token = %s,
                    refresh_token = %s
                WHERE name = %s
            """, (new_at, new_rt, doc_name))
            
            frappe.db.commit()
            return new_at
        else:
            frappe.log_error(f"Zalo trả về lỗi: {res_data}", "Zalo API Fail")
            return None

    except Exception:
        error_msg = traceback.format_exc()
        frappe.log_error(error_msg, "refresh_zalo_tokens SQL Failed")
        return None
    
def call_zalo_api(endpoint, method="GET", data=None):
    try:
        # Lấy access_token từ bảng riêng
        at_result = frappe.db.sql("""
            SELECT access_token 
            FROM `tabZalo Config` 
            LIMIT 1
        """)
        
        if not at_result:
            return {"error": -1, "message": "No Config Found"}
            
        at = at_result[0][0]
        headers = {'access_token': at}

        # Gọi API
        if method == "GET":
            response = requests.get(endpoint, headers=headers).json()
        else:
            response = requests.post(endpoint, headers=headers, json=data).json()

        # Nếu hết hạn thì Refresh
        if response.get("error") == -216:
            new_at = refresh_zalo_tokens()
            if new_at:
                headers['access_token'] = new_at
                # Thử lại lần 2
                if method == "GET":
                    response = requests.get(endpoint, headers=headers).json()
                else:
                    response = requests.post(endpoint, headers=headers, json=data).json()
                    
        return response

    except Exception:
        frappe.log_error(traceback.format_exc(), "call_zalo_api SQL Failed")
        return {"error": -1}

def send_zalo_vote_link_group(message):
    """
    Gửi thực đơn vào nhóm Zalo bằng GMF - Đã tích hợp tự động Refresh Token
    """
    url = "https://openapi.zalo.me/v3.0/oa/group/message"

    payload = {
        "recipient": {
            "group_id": GROUP_ID_ZALO  
        },
        "message": {
            "text": message
        }
    }

    try:
        response_data = call_zalo_api(url, method="POST", data=payload)
        
        if response_data.get("error") == 0:
            logger.info(f"[Zalo Group] Gửi thành công! Response: {response_data}")
        else:
            logger.error(f"[Zalo Group] Gửi thất bại: {response_data}")
            
    except Exception as e:
        logger.error(f"[Zalo Group] Lỗi hệ thống: {str(e)}")


@frappe.whitelist(allow_guest=True)
def check_and_renew_sessions():
    try:
        
        today_date = add_days(getdate(), 1)

        existing_today = frappe.db.sql("""
            SELECT name
            FROM `tabLunch Session`
            WHERE date = %s
            LIMIT 1
        """, today_date)

        if existing_today:
            logger.info("Session for today already exists -> EXIT")
            return

        # lấy session gần nhất
        last_session = frappe.db.sql("""
            SELECT *
            FROM `tabLunch Session`
            ORDER BY creation DESC
            LIMIT 1
        """, as_dict=True)

        if not last_session:
            logger.info("No Lunch Session found")
            return

        last_session = last_session[0]

        # check hết hạn
        expired = frappe.db.sql("""
            SELECT name
            FROM `tabLunch Session`
            WHERE name=%s
            AND status='Open'
            AND end_date < %s
        """, (last_session["name"] , today_date))

        if not expired:
            logger.info("Session not expired")
            return

        logger.info(f"Session expired: {last_session['name']}")

        # đóng session cũ
        frappe.db.sql("""
            UPDATE `tabLunch Session`
            SET status='Closed'
            WHERE name=%s
        """, last_session["name"])

        # tạo session mới
        new_name = frappe.generate_hash(10)
        link = f"{BASE_URL}/api/method/food_order_app.api.start_vote?session={new_name}"
        start_time = add_days(last_session["start_date"], 1)
        end_time = add_days(last_session["end_date"], 1)

        frappe.db.sql("""
            INSERT INTO `tabLunch Session`
            (
                name,
                session_name,
                date,
                start_date,
                end_date,
                status,
                vote_link,
                created_by,
                creation,
                modified
            )
            VALUES (%s,%s,%s,%s,%s,'Open',%s,'Administrator',%s,%s)
        """, (
            new_name,
            f"Menu {today_date}",
            today_date,
            start_time,
            end_time,
            link,
            today_date,today_date
        ))

        logger.info(f"Created new session: {new_name}")

        # 5️⃣ copy menu từ session cũ
        frappe.db.sql("""
            INSERT INTO `tabLunch Session Menu`
            (
                name,
                parent,
                parenttype,
                parentfield,
                menu_item,
                creation,
                modified
            )
            SELECT
                UUID(),
                %s,
                'Lunch Session',
                'menu_items',
                menu_item,
                %s,
                %s
            FROM `tabLunch Session Menu`
            WHERE parent=%s
        """, (new_name, today_date, today_date, last_session["name"]))

        frappe.db.commit()

        message = f"🔔 Đã có lịch đăng ký ăn trưa ngày {today_date}!\nKính mời anh/chị đăng ký tại:\n{link}"

        send_zalo_vote_link_group(message)

        logger.info("Database commit success")

    except Exception:

        error_trace = traceback.format_exc()

        logger.error(error_trace)

        frappe.log_error(
            title="check_and_renew_sessions failed",
            message=error_trace
        )

        frappe.db.rollback()

@frappe.whitelist(allow_guest=True)
def remind_vote_today():
    """
    Gửi tin nhắc lúc 9h sáng: 
    Đã có X người chọn món rồi, đừng quên bình chọn nhé!
    """
    try:
        today_date = getdate()

        session = frappe.db.sql("""
            SELECT name, vote_link
            FROM `tabLunch Session`
            WHERE date=%s AND status='Open'
            LIMIT 1
        """, today_date, as_dict=True)

        if not session:
            logger.info("No open session for today -> EXIT")
            return

        session = session[0] 
        session_name = session["name"]
        vote_link = session["vote_link"]

        vote_count = frappe.db.sql("""
            SELECT COUNT(*) 
            FROM `tabLunch Order`
            WHERE session=%s AND is_active=1
        """, session_name)[0][0]

        message = (
            f"⏰ Nhắc hẹn đăng ký bữa trưa ngày {today_date}!\n"
            f"Hiện tại đã có {vote_count} người đăng ký bữa rồi.\n"
            f"Đừng quên đăng ký tại đây nhé!\n"
            f"{vote_link}"
        )

        # gửi group zalo
        send_zalo_vote_link_group(message)

        logger.info(f"Sent reminder message for session {session_name}")

    except Exception:
        error = traceback.format_exc()
        logger.error(error)
        frappe.log_error("remind_vote_today failed", error)


@frappe.whitelist(allow_guest=False)
def export_monthly_report(month=None, year=None):
    """
    Export Excel Report cho Lunch Session
    Method: GET /api/method/food_order_app.api.export_monthly_report?month=3&year=2026
    """
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
    from io import BytesIO
    import calendar
    from datetime import datetime
    
    if not month or not year:
        now = datetime.now()
        month = now.month
        year = now.year
        
    month = int(month)
    year = int(year)
    
    # 1. Tính số ngày trong tháng
    _, days_in_month = calendar.monthrange(year, month)
    
    # 2. Lấy dữ liệu Order
    orders = frappe.db.sql("""
        SELECT 
            lo.zalo_user,
            DAY(ls.date) as order_day,
            lmi.price
        FROM `tabLunch Order` lo
        JOIN `tabLunch Session` ls ON lo.session = ls.name
        JOIN `tabLunch Menu Item` lmi ON lo.menu_item = lmi.name
        WHERE MONTH(ls.date) = %s AND YEAR(ls.date) = %s AND ls.status != 'Draft'
    """, (month, year), as_dict=True)
    
    user_orders = {}
    for o in orders:
        zalo_user = o.zalo_user
        if zalo_user not in user_orders:
            user_orders[zalo_user] = {'days': set(), 'total_amount': 0}
        user_orders[zalo_user]['days'].add(o.order_day)
        user_orders[zalo_user]['total_amount'] += (o.price or 0)
        
    # 3. Lấy dữ liệu Ví
    wallets = frappe.get_all("Lunch Wallet", fields=["zalo_user", "balance"])
    wallet_map = {w.zalo_user: w.balance for w in wallets}
    
    # 4. Tạo file Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Tháng {month}-{year}"
    
    header_fill = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")
    header_font = Font(bold=True)
    center_align = Alignment(horizontal='center', vertical='center')
    
    # Dòng 1 (Merge ô hiển thị "Ngày")
    ws.merge_cells(start_row=1, start_column=3, end_row=1, end_column=2+days_in_month)
    day_cell = ws.cell(row=1, column=3, value="Ngày")
    day_cell.alignment = center_align
    day_cell.font = header_font
    day_cell.fill = header_fill
    
    # Style phần còn lại của dòng 1
    for col in range(1, 3+days_in_month+4):
        c = ws.cell(row=1, column=col)
        c.fill = header_fill
    
    # Dòng 2 (Tiêu đề cụ thể)
    headers = ["STT", "Họ và tên"] + list(range(1, days_in_month + 1)) + [
        "Số ngày ăn", "Đơn giá suất ăn \n(VNĐ)", "Thành tiền \n(VNĐ)", "Số tiền còn lại \n(VNĐ)"
    ]
    ws.append(headers)
    
    # Style dòng 2
    for col_idx, value in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col_idx)
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.fill = header_fill
        
        # Căn chỉnh độ rộng cột
        if col_idx == 1:
            ws.column_dimensions[get_column_letter(col_idx)].width = 5
        elif col_idx == 2:
            ws.column_dimensions[get_column_letter(col_idx)].width = 25
        elif col_idx > 2 and col_idx <= 2 + days_in_month:
            ws.column_dimensions[get_column_letter(col_idx)].width = 4
        else:
            ws.column_dimensions[get_column_letter(col_idx)].width = 15
            
    # 5. Đổ dữ liệu
    users = frappe.get_all("Zalo User Map", fields=["name", "full_name"])
    stt = 1
    
    for u in users:
        u_data = user_orders.get(u.name, {'days': set(), 'total_amount': 0})
        num_days = len(u_data['days'])
        total_price = u_data['total_amount']
        avg_price = (total_price / num_days) if num_days > 0 else 0
        balance = wallet_map.get(u.name, 0)
        
        row_data = [stt, u.full_name]
        
        # Cột ngày (1-31)
        for d in range(1, days_in_month + 1):
            row_data.append(1 if d in u_data['days'] else "")
            
        # Các cột tổng hợp
        row_data.extend([
            num_days,
            avg_price,
            total_price,
            balance
        ])
        
        ws.append(row_data)
        
        # Style row data
        curr_row = ws.max_row
        for col_idx in range(1, len(row_data) + 1):
            cell = ws.cell(row=curr_row, column=col_idx)
            if col_idx > 2: # Can_center số
                cell.alignment = center_align
            # Format tiền tệ
            if col_idx in [len(row_data) - 2, len(row_data) - 1, len(row_data)]:
                cell.number_format = '#,##0'
                
        stt += 1
        
    # Cố định header
    ws.freeze_panes = "C3"

    # Trả về File    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    frappe.response['filename'] = f"BaoCao_AnTrua_Thang_{month}_{year}.xlsx"
    frappe.response['filecontent'] = output.getvalue()
    frappe.response['type'] = 'binary'
